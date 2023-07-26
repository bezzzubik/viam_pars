import pandas
import warnings
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.axis import DateAxis


import json
from pymysql import connect, OperationalError
import os
from string import Template

"""классы и функции для работы с бд"""


class SQLProvider:
    def __init__(self, file_path: str) -> None:
        self._scripts = {}
        for file in os.listdir(file_path):  # file_path - путь до папки sql
            self._scripts[file] = Template(
                open(f'{file_path}/{file}', 'r').read())

    def get(self, name, **kwargs):  # get('avto_by_lift.sql', lift=...
        return self._scripts[name].substitute(**kwargs)


class DBConnection:
    def __init__(self, config):
        self.config = config
        self.connection = None
        self.cursor = None

    def __enter__(self):
        try:
            self.connection = connect(**self.config)
            self.cursor = self.connection.cursor()
            return self.cursor
        except OperationalError:
            return None

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.connection is not None and self.cursor is not None:
            self.connection.commit()
            self.connection.close()
            self.cursor.close()
        if exc_val is not None:
            print(exc_type)
            print(exc_val.args)
        return True


def work_with_db(config, sql) -> list:
    result = []
    with DBConnection(config) as cursor:
        cursor.execute(sql)
        schema = [column[0] for column in cursor.description]
        for item in cursor.fetchall():
            result.append(dict(zip(schema, item)))
    return result


def db_update(config: dict, _sql):
    with DBConnection(config) as cursor:
        if cursor is None:
            raise ValueError('Курсор None')
        elif cursor:
            cursor.execute(_sql)

provider = SQLProvider(os.path.join(os.path.dirname('main.  py'), 'sql'))

conf = json.load(open('configs/sql_config.json'))

"""классы и функции для работы с бд"""

"""Основная программа"""

wb = Workbook()
wb.create_sheet(title='Первый лист', index=0)
sheet = wb['Первый лист']
sheet['A1'] = 'Неделя'
sheet['B1'] = 'Накопившиеся полученные образцы'
sheet['D1'] = 'Испытанные образцы за неделю'
sheet['C1'] = 'Накопившиеся испытанные образцы'
mon={1:'январь',
     2:'февраль',
     3:'март',
     4:'апрель',
     5:'май',
     6:'июнь',
     7:'июль',
     8:'август',
     9:'сентябрь',
     10:'октябрь',
     11:'ноябрь',
     12:'декабрь'}
j=0

weeks = work_with_db(conf, provider.get('select_min_max_week.sql'))
minW=weeks[0]['min']
maxW=weeks[0]['max']

sumIn=0
sumEnd=0
weeks=[]
for i in range(minW, maxW+1):
    countI = work_with_db(conf, provider.get('select_sum_factIn.sql', weekS=i))
    countE = work_with_db(conf, provider.get('select_sum_factEnd.sql', weekS=i))
    monNow = work_with_db(conf, provider.get('select_min_week_in_month.sql', mn=j+1))

    countI = countI[0]['factIn']
    countE = countE[0]['factEnd']
    monNow = monNow[0]['mdat']

    if countI is None:
        countI = 0    
    if countE is None:
        countE = 0
    sumIn += countI
    sumEnd += countE
    print(i)
    if monNow == i:
        j+=1
        wk=mon[j]
    else:
        wk=i
    sheet.cell(row = 2+i-minW, column=1).value=wk
    sheet.cell(row = 2+i-minW, column=2).value=sumIn
    sheet.cell(row = 2+i-minW, column=3).value=sumEnd
    sheet.cell(row = 2+i-minW, column=4).value=countE

print('**********************')
print(sumIn)
print(sumEnd)
print('**********************')

chartBar = LineChart()
chartBar.y_axis.title='Кол-во образцов'
chartBar.x_axis.title='Номер недели'
chartBar.title = 'Графики'
chartBar.style = 12
chartBar.width = 35
chartBar.height = 20

dataBar = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=maxW)
dataX = Reference(sheet, min_col=1, min_row=2, max_col=1, max_row=maxW)

chartBar.add_data(dataBar, titles_from_data=True)
chartBar.smooth = True
chartBar.set_categories(dataX)
s1=chartBar.series[1]
s1.graphicalProperties.line.solidFill = "008000"



s3=chartBar.series[2]
s3.graphicalProperties.line.solidFill = "00AAAA"
s3.graphicalProperties.line.dashStyle = "sysDot"
s3.graphicalProperties.line.width = 100050 # width in EMUs


sheet.add_chart(chartBar, 'G2')
wb.save('График.xlsx')